# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from functools import wraps
from django.http import HttpResponse, HttpResponseNotFound
from django.conf import settings
from django.core import signing
from django.core.files import File
from django.core.files.storage import FileSystemStorage, default_storage
import os
import io
import re
import xlsxwriter
import openpyxl

from .models import AdminTable, Customer, Container, Order
from .utils import generate_password, send_credentials_email, encrypt_text, decrypt_text
from django.db.models import Sum

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT

# ----------------------------
# Decorators
# ----------------------------
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get('admin_id'):
            return redirect('admin_login')
        return view_func(request, *args, **kwargs)
    return wrapper

def customer_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.session.get('customer_id'):
            return redirect('user_login')
        return view_func(request, *args, **kwargs)
    return wrapper


# ----------------------------
# NEW START
# ----------------------------


def manage_customer_view(request):
    # Your existing view logic for handling forms and getting customers...
    customers = ... # Query your customers
    
    context = {
        'customers': customers,
        'active_page': 'customers'  # This is the key for the sidebar
    }
    return render(request, 'manage_customer.html', context)

def manage_container_view(request):
    # Your existing view logic for handling forms and getting containers...
    containers = ... # Query your containers

    context = {
        'containers': containers,
        'active_page': 'containers' # This is the key for the sidebar
    }
    return render(request, 'manage_container.html', context)


def dashboard_view(request):
    context = {
        'active_page': 'dashboard' 
    }
    return render(request, 'dashboard.html', context)


def add_order_view(request):
    # Your logic to handle file upload, read excel, generate preview_data

    containers = Container.objects.all() # Fetch all containers
    preview_data = request.session.get('preview_data', None) # Get preview data if it exists

    context = {
        'active_page': 'add_order', # For highlighting the sidebar link
        'containers': containers,
        'selected_container': request.POST.get('container', None), # Keep dropdown selected
        'preview_data': preview_data
    }
    return render(request, 'add_order.html', context)

def view_order_view(request):
    containers = Container.objects.all()
    selected_container_id = request.GET.get('container', None)
    orders = None

    if selected_container_id:
        orders = Order.objects.filter(container__container_id=selected_container_id)

    context = {
        'active_page': 'view_order', # For the sidebar
        'containers': containers,
        'selected_container_id': selected_container_id,
        'orders': orders
    }
    return render(request, 'view_order.html', context)

def view_customer_view(request):
    all_customers = Customer.objects.all()
    selected_customer = None
    customer_stats = None
    recent_orders = None

    customer_id = request.GET.get('customer')
    if customer_id:
        try:
            selected_customer = Customer.objects.get(id=customer_id)
            
            # Calculate stats
            orders = Order.objects.filter(customer=selected_customer)
            customer_stats = {
                'total_orders': orders.count(),
                'active_shipments': orders.filter(status__in=['in_transit', 'at_port']).count(),
                'total_value': orders.aggregate(Sum('value'))['value__sum'] or 0,
            }

            # Fetch recent orders
            recent_orders = orders.order_by('-order_date')[:5] # Get latest 5 orders

        except Customer.DoesNotExist:
            # Handle case where customer is not found
            pass

    context = {
        'active_page': 'view_customer',
        'all_customers': all_customers,
        'selected_customer': selected_customer,
        'customer_stats': customer_stats,
        'recent_orders': recent_orders,
    }
    return render(request, 'view_customer.html', context)



def field_visibility_view(request):
    if request.method == 'POST':
        for field in FieldVisibility.objects.all():
            # The value will be 'on' if checked, None if not.
            is_visible = request.POST.get(field.field_name) == 'on'
            if field.is_visible != is_visible:
                field.is_visible = is_visible
                field.save()
        messages.success(request, 'Visibility settings updated successfully.')
        return redirect('field_visibility')

    # For GET request
    fields = FieldVisibility.objects.all().order_by('field_name')
    context = {
        'active_page': 'field_visibility',
        'fields': fields
    }
    return render(request, 'field_visibility.html', context)

# ----------------------------
# NEW END
# ----------------------------


# ----------------------------
# Home
# ----------------------------
def home(request):
    return render(request, "home.html")

# ----------------------------
# Admin Views
# ----------------------------
def admin_login(request):
    if request.method == "POST":
        user_id = request.POST.get("username")
        password = request.POST.get("password")

        try:
            admin = AdminTable.objects.get(user_id=user_id)
            if admin.check_password(password):
                request.session['admin_id'] = admin.admin_id
                request.session['admin_user_id'] = admin.user_id
                return redirect('admin_dashboard')
            else:
                messages.error(request, "Invalid password")
        except AdminTable.DoesNotExist:
            messages.error(request, "Admin not found")

    return render(request, "admin_login.html")

@admin_required
def admin_dashboard(request):
    admin_user_id = request.session.get('admin_user_id')
    return render(request, "admin_dashboard.html", {"admin_user_id": admin_user_id})

@admin_required
def admin_containers(request):
    containers = Container.objects.all()

    if request.method == "POST":
        container_no = request.POST.get("container_no")
        container_name = request.POST.get("container_name")
        if container_no and container_name:
            container = Container(container_no=container_no, container_name=container_name)
            container.save()
            messages.success(request, f"Container {container.container_id} added successfully")
            return redirect('admin_containers')

    return render(request, "admin_containers.html", {"containers": containers})

@admin_required
def delete_container(request, container_id):
    container = get_object_or_404(Container, container_id=container_id)

    # Get all related orders
    related_orders = Order.objects.filter(container=container)
    order_count = related_orders.count()

    # Delete images from MEDIA_ROOT
    for order in related_orders:
        if order.pictures and os.path.isfile(order.pictures.path):
            try:
                os.remove(order.pictures.path)
            except Exception as e:
                print(f"Error deleting image {order.pictures.path}: {e}")

    # Delete related orders
    related_orders.delete()

    # Now delete the container
    container.delete()

    messages.success(
        request,
        f"Container {container_id} and its {order_count} related order(s) deleted successfully, including images."
    )
    return redirect('admin_containers')
# from .decorators import admin_required

@admin_required
def admin_customers(request):
    import pandas as pd
    from .utils import generate_password, send_credentials_email
    from django.contrib import messages

    customers = Customer.objects.all()

    if request.method == "POST":
        # Check if XLS upload
        if "upload_xls" in request.FILES:
            xls_file = request.FILES['upload_xls']
            df = pd.read_excel(xls_file)

            for _, row in df.iterrows():
                user_id = row['user_id']
                if Customer.objects.filter(user_id=user_id).exists():
                    messages.warning(request, f"User ID {user_id} skipped: already exists.")
                    continue

                password = generate_password()
                customer = Customer(
                    customer_id=row['customer_id'],
                    name=row['name'],
                    user_id=row['user_id'],
                    email=row['email'],
                    whatsapp_number=row.get('whatsapp_number', '')
                )
                customer.set_password(password)
                customer.save()
                send_credentials_email(row['email'], user_id, password)
                messages.success(request, f"User ID {user_id} added successfully.")

        else:
            # Single customer add
            customer_id = request.POST.get("customer_id")
            name = request.POST.get("name")
            user_id = request.POST.get("user_id")
            email = request.POST.get("email")
            whatsapp = request.POST.get("whatsapp_number")

            if Customer.objects.filter(user_id=user_id).exists():
                messages.error(request, "User ID already exists")
            else:
                password = generate_password()
                customer = Customer(
                    customer_id=customer_id,
                    name=name,
                    user_id=user_id,
                    email=email,
                    whatsapp_number=whatsapp
                )
                customer.set_password(password)
                customer.save()
                send_credentials_email(email, user_id, password)
                messages.success(request, f"Customer {customer_id} added successfully")

        return redirect('admin_customers')

    return render(request, "admin_customers.html", {"customers": customers})


@admin_required
def admin_logout(request):
    request.session.flush()
    messages.success(request, "Admin logged out successfully")
    return redirect('admin_login')


@admin_required
def add_order(request):
    customers = Customer.objects.all()
    containers = Container.objects.all()

    if request.method == "POST":
        customer_id = request.POST.get("customer")
        container_id = request.POST.get("container")
        shipping_mark = request.POST.get("shipping_mark")
        description = request.POST.get("description")
        item_no_spec = request.POST.get("item_no_spec")
        material = request.POST.get("material")
        pictures = request.FILES.get("pictures")

        # Numeric fields with proper casting
        ctns = int(request.POST.get("ctns", 0))
        qty_per_ctn = int(request.POST.get("qty_per_ctn", 0))
        total_qty = ctns * qty_per_ctn

        unit = request.POST.get("unit")
        cbm_per_ctn = float(request.POST.get("cbm_per_ctn", 0))
        cbm = float(request.POST.get("cbm", 0))
        total_cbm = float(request.POST.get("total_cbm", 0))
        wt_per_ctn = float(request.POST.get("wt_per_ctn", 0))
        total_wt = float(request.POST.get("total_wt", 0))
        supplier = request.POST.get("supplier")

        customer = get_object_or_404(Customer, customer_id=customer_id)
        container = get_object_or_404(Container, container_id=container_id)

        Order.objects.create(
            customer=customer,
            container=container,
            shipping_mark=shipping_mark,
            description=description,
            item_no_spec=item_no_spec,
            material=material,
            pictures=pictures,
            ctns=ctns,
            qty_per_ctn=qty_per_ctn,
            total_qty=total_qty,
            unit=unit,
            cbm_per_ctn=cbm_per_ctn,
            cbm=cbm,
            total_cbm=total_cbm,
            wt_per_ctn=wt_per_ctn,
            total_wt=total_wt,
            supplier=supplier
        )

        messages.success(request, f"Order added successfully for {customer.customer_id}")
        return redirect("add_order")

    return render(request, "add_order.html", {"customers": customers, "containers": containers})


import os
import re
import openpyxl
from django.core.files import File
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.conf import settings

from .models import Container, Customer, Order
# from .decorators import admin_required


@admin_required
def upload_orders_excel(request):
    containers = Container.objects.all()
    preview_data = None
    selected_container = None
    temp_images = set()

    if request.method == "POST" and request.FILES.get("excel_file"):
        selected_container = request.POST.get("container")
        excel_file = request.FILES["excel_file"]

        # Save uploaded Excel temporarily
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, "excel_uploads"))
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active

            # Save merged ranges
            merged_ranges = list(sheet.merged_cells.ranges)
            for merged_range in merged_ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                first_val = sheet.cell(row=min_row, column=min_col).value
                sheet.unmerge_cells(range_string=str(merged_range))
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        sheet.cell(row=r, column=c, value=first_val)

            # Extract images to media/order_pics
            images_dir = os.path.join(settings.MEDIA_ROOT, "order_pics")
            os.makedirs(images_dir, exist_ok=True)
            images_map = {}

            for img in sheet._images:
                try:
                    img_bytes = img._data()
                    img_filename = f"img_{len(temp_images)+1}.png"
                    img_path = os.path.join(images_dir, img_filename)
                    with open(img_path, "wb") as f:
                        f.write(img_bytes)

                    # Anchor row/col
                    anchor_row = img.anchor._from.row + 1
                    anchor_col = img.anchor._from.col + 1
                    assigned = False

                    for merged_range in merged_ranges:
                        min_col, min_row, max_col, max_row = merged_range.bounds
                        if min_row <= anchor_row <= max_row and min_col <= anchor_col <= max_col:
                            for r in range(min_row, max_row + 1):
                                images_map.setdefault(r, set()).add(f"order_pics/{img_filename}")
                                temp_images.add(f"order_pics/{img_filename}")
                            assigned = True
                            break

                    if not assigned:
                        images_map.setdefault(anchor_row, set()).add(f"order_pics/{img_filename}")
                        temp_images.add(f"order_pics/{img_filename}")

                except Exception as e:
                    print("Image extraction failed:", e)

            # Normalize headers (row 3)
            headers = []
            for cell in sheet[3]:
                if cell.value:
                    header = str(cell.value).replace("\n", " ").replace('"', '').strip()
                    if re.search(r"shipping\s*mark", header, re.I):
                        header = "Shipping Mark"
                    elif re.search(r"item\s*no.*spec", header, re.I):
                        header = "Item No. /Specification"
                    elif re.search(r"total\s*qty", header, re.I):
                        header = "Total Qty"
                    headers.append(header)
                else:
                    headers.append(f"Column{len(headers)+1}")

            # Collect rows
            raw_rows = []
            for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
                if all(v is None for v in row):
                    continue
                row_dict = dict(zip(headers, row))
                if i in images_map:
                    row_dict["Pictures"] = list(images_map[i])
                raw_rows.append(row_dict)

            # Group images by (Customer, Shipping Mark)
            preview_data = []
            customer_shipping_images = {}
            for row_dict in raw_rows:
                customer = str(row_dict.get("CUSTOMER")).strip() if row_dict.get("CUSTOMER") else None
                shipping_mark = str(row_dict.get("Shipping Mark")).strip() if row_dict.get("Shipping Mark") else None

                if "Pictures" in row_dict and row_dict["Pictures"]:
                    key = (customer, shipping_mark)
                    customer_shipping_images.setdefault(key, set()).update(row_dict["Pictures"])

                preview_data.append(row_dict)

            # Attach grouped images
            for row_dict in preview_data:
                customer = str(row_dict.get("CUSTOMER")).strip() if row_dict.get("CUSTOMER") else None
                shipping_mark = str(row_dict.get("Shipping Mark")).strip() if row_dict.get("Shipping Mark") else None

                key = (customer, shipping_mark)
                if ("Pictures" not in row_dict or not row_dict.get("Pictures")) and key in customer_shipping_images:
                    row_dict["Pictures"] = list(customer_shipping_images[key])

            # Save preview and temp images in session (convert sets to lists!)
            request.session["excel_preview"] = preview_data
            request.session["selected_container"] = selected_container
            request.session["temp_order_images"] = list(temp_images)

        except Exception as e:
            messages.error(request, f"Error processing Excel file: {e}")
            preview_data = None

        finally:
            if wb:
                wb.close()
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except PermissionError:
                print(f"Could not delete {file_path}, still in use.")

    return render(request, "upload_orders_excel.html", {
        "containers": containers,
        "preview_data": preview_data,
        "selected_container": selected_container,
        "MEDIA_URL": settings.MEDIA_URL,
    })


@admin_required
def confirm_orders_excel(request):
    preview_data = request.session.get("excel_preview")
    selected_container = request.session.get("selected_container")
    temp_images = set(request.session.get("temp_order_images", []))  # Convert back to set

    if not preview_data or not selected_container:
        messages.error(request, "No preview data found. Please upload again.")
        return redirect("upload_orders_excel")

    container = get_object_or_404(Container, container_id=selected_container)
    missing_customers = []

    for row in preview_data:
        customer_id = str(row.get("CUSTOMER")).strip() if row.get("CUSTOMER") else None
        if not customer_id:
            continue

        customer = Customer.objects.filter(customer_id=customer_id).first()
        if not customer:
            missing_customers.append(customer_id)
            continue

        try:
            order = Order(
                customer=customer,
                container=container,
                shipping_mark=row.get("Shipping Mark") or None,
                description=row.get("Description") or None,
                item_no_spec=row.get("Item No. /Specification") or None,
                material=row.get("Material") or None,
                ctns=row.get("Ctns") or None,
                qty_per_ctn=row.get("Qty/Ctn") or None,
                total_qty=row.get("Total Qty") or None,
                unit=row.get("Unit") or None,
                cbm_per_ctn=row.get("CBM/Ctn") or None,
                cbm=row.get("CBM") or None,
                total_cbm=row.get("Total CBM") or None,
                wt_per_ctn=row.get("Wt/Ctn") or None,
                total_wt=row.get("Total Wt") or None,
                supplier=row.get("Supplier") or None,
            )

            # Attach images
            if row.get("Pictures"):
                for img_relative_path in row["Pictures"]:
                    img_path = os.path.join(settings.MEDIA_ROOT, img_relative_path)
                    if os.path.exists(img_path) and not order.pictures:
                        order.pictures.save(
                            os.path.basename(img_path),
                            File(open(img_path, "rb")),
                            save=False
                        )

            order.save()

        except Exception as e:
            messages.warning(request, f"Error saving order for Customer {customer_id}: {e}")

    # Delete temporary images after saving
    for img_relative_path in temp_images:
        img_path = os.path.join(settings.MEDIA_ROOT, img_relative_path)
        if os.path.exists(img_path):
            try:
                os.remove(img_path)
            except Exception as e:
                print(f"Could not delete temporary image {img_path}: {e}")

    # Clear session
    for key in ["excel_preview", "selected_container", "temp_order_images"]:
        request.session.pop(key, None)

    if missing_customers:
        messages.warning(request, f"The following customer IDs were not found in database: {', '.join(missing_customers)}")

    messages.success(request, "Orders imported successfully!")
    return redirect("admin_dashboard")

# view_orders_by_container

@admin_required
def view_orders_by_container(request):
    containers = Container.objects.all()
    selected_container_id = request.GET.get("container")
    orders = []

    if selected_container_id:
        container = get_object_or_404(Container, container_id=selected_container_id)
        orders = Order.objects.filter(container=container).order_by("customer__customer_id")

    return render(request, "view_orders_by_container.html", {
        "containers": containers,
        "selected_container_id": selected_container_id,
        "orders": orders,
    })


@admin_required
def view_orders_by_customer(request):
    containers = Container.objects.all()
    selected_container_id = None
    selected_customer_id = None
    customers = []
    orders = []

    # ----- Decode container from GET -----
    container_signed = request.GET.get("container")
    if container_signed:
        try:
            selected_container_id = signing.loads(container_signed)
        except signing.BadSignature:
            selected_container_id = None

    # ----- Decode customer from GET -----
    customer_signed = request.GET.get("customer")
    if customer_signed:
        try:
            decoded = signing.loads(customer_signed)
            selected_container_id = decoded.get("container")
            selected_customer_id = decoded.get("customer")
        except signing.BadSignature:
            selected_customer_id = None

    container = None
    if selected_container_id:
        container = get_object_or_404(Container, container_id=selected_container_id)

        # Populate customer dropdown for selected container
        customers_raw = Order.objects.filter(container=container).values_list('customer__customer_id', flat=True).distinct()
        customers = []
        for c in customers_raw:
            signed_data = signing.dumps({'container': selected_container_id, 'customer': c})
            customers.append({'id': c, 'signed': signed_data})

    # Fetch orders only if both container and customer are selected
    if selected_container_id and selected_customer_id:
        orders = Order.objects.filter(
            container__container_id=selected_container_id,
            customer__customer_id=selected_customer_id
        )

    # Sign container IDs for dropdown
    signed_containers = []
    for c in containers:
        signed_containers.append({
            'id': c.container_id,
            'signed': signing.dumps(c.container_id)
        })

    return render(request, "view_orders_by_customer.html", {
        "containers": signed_containers,
        "selected_container_id": selected_container_id,
        "customers": customers,
        "selected_customer_id": selected_customer_id,
        "orders": orders,
    })




# views.py

# from django.contrib.admin.views.decorators import staff_member_required
# views.py
from django.shortcuts import render
from django.contrib import messages
from .models import OrderFieldVisibility
# from .decorators import admin_required

@admin_required
def manage_field_visibility(request):
    # Ensure all fields exist in DB
    fields_in_db = {f.field_name: f for f in OrderFieldVisibility.objects.all()}
    all_fields = []

    for field_name, field_label in OrderFieldVisibility.FIELD_CHOICES:
        if field_name in fields_in_db:
            field_obj = fields_in_db[field_name]
        else:
            # create default record if not exists
            field_obj = OrderFieldVisibility.objects.create(field_name=field_name, is_visible=True)
        all_fields.append(field_obj)

    if request.method == "POST":
        for field in all_fields:
            field.is_visible = request.POST.get(field.field_name) == "on"
            field.save()
        messages.success(request, "Field visibility updated successfully.")

    return render(request, "manage_field_visibility.html", {"fields": all_fields})


# ----------------------------
# Customer Views
# ----------------------------
def user_login(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        password = request.POST.get("password")

        try:
            customer = Customer.objects.get(user_id=user_id)
            if customer.check_password(password):
                # Set session
                request.session['customer_id'] = customer.customer_id
                request.session['customer_name'] = customer.name

                # First time login check
                if customer.is_first_login:
                    return redirect('user_update_password')
                return redirect('user_dashboard')
            else:
                messages.error(request, "Invalid password")
        except Customer.DoesNotExist:
            messages.error(request, "User not found")

    return render(request, "user_login.html")


@customer_required
def user_dashboard(request):
    customer_id = request.session.get('customer_id')
    customer = get_object_or_404(Customer, customer_id=customer_id)
    return render(request, "user_dashboard.html", {"customer": customer})


@customer_required
def user_update_password(request):
    customer_id = request.session.get('customer_id')
    customer = get_object_or_404(Customer, customer_id=customer_id)

    if request.method == "POST":
        new_password = request.POST.get("password")
        if not new_password:
            messages.error(request, "Password cannot be empty")
        else:
            customer.set_password(new_password)  # Hashes the password
            customer.is_first_login = False
            customer.save()
            messages.success(request, "Password updated successfully")
            return redirect('user_dashboard')

    return render(request, "user_update_password.html", {"customer": customer})


@customer_required
def user_logout(request):
    request.session.flush()
    messages.success(request, "User logged out successfully")
    return redirect('user_login')


# ----------------------------
# List all containers for the logged-in user
# ----------------------------
# from .utils import encrypt_text  # assuming you have this utility

def get_visible_fields():
    return {fv.field_name for fv in OrderFieldVisibility.objects.filter(is_visible=True)}


@customer_required
def user_containers(request):
    customer_id = request.session.get('customer_id')

    # Get unique containers for this customer
    containers = Container.objects.filter(order__customer_id=customer_id).distinct()

    # Prepare list of dicts with container and encrypted ID
    container_links = []
    for c in containers:
        container_links.append({
            "container": c,
            "enc_id": encrypt_text(c.container_id)
        })

    return render(request, "user_containers.html", {
        "container_links": container_links
    })

# ----------------------------
# Show orders inside a selected container (with images)
# ----------------------------
@customer_required
def user_container_orders(request, enc_container_id):
    customer_id = request.session.get('customer_id')
    container_id = decrypt_text(enc_container_id)
    container = get_object_or_404(Container, container_id=container_id)

    orders = Order.objects.filter(container=container, customer_id=customer_id)
    total_qty_sum = orders.aggregate(total=Sum('total_qty'))['total'] or 0

    visible_fields = get_visible_fields()

    return render(request, "user_container_orders.html", {
        "container": container,
        "orders": orders,
        "total_qty_sum": total_qty_sum,
        "enc_container_id": enc_container_id,
        "visible_fields": visible_fields,
    })

# ----------------------------
# Export orders to PDF (with images)
# ----------------------------
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import io
import os
from django.conf import settings

@customer_required
def export_orders_pdf(request, enc_container_id):
    customer_id = request.session.get("customer_id")
    container_id = decrypt_text(enc_container_id)
    container = get_object_or_404(Container, container_id=container_id)
    orders = Order.objects.filter(container=container, customer_id=customer_id)

    visible_fields = set(
        OrderFieldVisibility.objects.filter(is_visible=True)
        .values_list("field_name", flat=True)
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="orders_{container.container_name}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Header labels
    field_labels = dict(OrderFieldVisibility.FIELD_CHOICES)
    headers = ["#"] + [field_labels[f] for f in field_labels if f in visible_fields]
    data = [headers]

    for idx, order in enumerate(orders, start=1):
        row = [str(idx)]
        for field in field_labels:
            if field not in visible_fields:
                continue
            value = getattr(order, field)
            
            # Handle image
            if field == "pictures" and value:
                img_path = os.path.join(settings.MEDIA_ROOT, value.name)
                if os.path.exists(img_path):
                    # Resize image to fit cell
                    img = Image(img_path, width=60, height=60)  # adjust size
                    row.append(img)
                else:
                    row.append("-")
            else:
                row.append(str(value) if value else "-")
        data.append(row)

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))

    elements.append(Paragraph(f"Orders in {container.container_name}", styles["Heading2"]))
    elements.append(table)
    doc.build(elements)

    return response


import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Container, Order, OrderFieldVisibility

@customer_required
def export_orders_excel(request, enc_container_id):
    customer_id = request.session.get("customer_id")
    container_id = decrypt_text(enc_container_id)
    container = get_object_or_404(Container, container_id=container_id)
    orders = Order.objects.filter(container=container, customer_id=customer_id)

    # Only fields marked visible
    visible_fields = set(
        OrderFieldVisibility.objects.filter(is_visible=True)
        .values_list("field_name", flat=True)
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    field_labels = dict(OrderFieldVisibility.FIELD_CHOICES)
    headers = ["#"] + [field_labels[f] for f in field_labels if f in visible_fields]
    ws.append(headers)

    # Data rows
    for idx, order in enumerate(orders, start=1):
        row = [idx]
        for field in field_labels:
            if field not in visible_fields:
                continue
            value = getattr(order, field)
            if field == "pictures" and value:
                row.append(value.url)
            else:
                row.append(str(value) if value else "")
        ws.append(row)

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 3

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="orders_{container.container_name}.xlsx"'
    return response


# ----------------------------
# Update Customer Profile
# ----------------------------
@customer_required
def user_update_profile(request):
    customer_id = request.session.get('customer_id')
    customer = get_object_or_404(Customer, customer_id=customer_id)

    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        whatsapp = request.POST.get("whatsapp_number")

        # Simple validation
        if not name or not email:
            messages.error(request, "Name and Email cannot be empty")
        else:
            customer.name = name
            customer.email = email
            customer.whatsapp_number = whatsapp
            customer.save()
            request.session['customer_name'] = customer.name  # Update session
            messages.success(request, "Profile updated successfully")
            return redirect('user_dashboard')

    return render(request, "user_update_profile.html", {"customer": customer})




def custom_404_view(request, exception=None):
    # If not logged in → redirect to login
    if not request.session.get("admin_id") and not request.session.get("customer_id"):
        return redirect("user_login")

    # If logged in → redirect to correct dashboard
    if request.session.get("admin_id"):
        return redirect("admin_dashboard")
    elif request.session.get("customer_id"):
        return redirect("user_dashboard")

    # fallback → show friendly 404 page
    return render(request, "404.html", status=404)


def custom_500_view(request):
    return render(request, "500.html", status=500)
